import os
import re
import math
import google_auth_oauthlib.flow
import googleapiclient.discovery
import googleapiclient.errors
from openpyxl import Workbook

# PLAYLISTS FOR TESTING
# PLc_ayA89mOhQgknIyASnCIwFAvMStUmSV smol playlist

PLAYLIST_ID = "PLc_ayA89mOhQgknIyASnCIwFAvMStUmSV"
MAX_PAGES = 100000

scopes = ["https://www.googleapis.com/auth/youtube.readonly"]
API_KEY = "put your key here"
API_SERVICE = "youtube"
API_VERSION = "v3"
YT_VID_PREFIX = "https://www.youtube.com/watch?v="
NND_VID_PREFIX = "http://www.nicovideo.jp/watch/"

TEST_LINK_1 = " http://www.nicovideo.jp/watch/sm18350640"
TEST_LINK_2 = "asdasd sm20342896"
TEST_LINK_3 = "asdasd http://nico.ms/sm33193478"

def main():
    os.environ["OAUTHLIB_INSECURE_TRANSPORT"] = "1"

    # regex to search for Nico Nico Douga links within video descriptions. Supports both https links and old http links, shortened url and just the video id
    # nndRegex = re.compile("(http(s)?://(www\\.nicovideo\\.jp/watch/sm[0-9]{8})|(nico\\.ms/sm[0-9]{8}))|(sm[0-9]{8})")
    
    # well this is much easier lol
    nndRegex = re.compile("(sm[0-9]{8})")

    # just some code to test regex quickly
    # nnd = nndRegex.search(TEST_LINK_3)
    # if (nnd != None):
    #     nndLink = nnd.group(0)
    #     print(nndLink)

    youtube = googleapiclient.discovery.build(
        API_SERVICE, API_VERSION, developerKey=API_KEY)

    # request playlist info
    response = youtube.playlists().list(
        part="snippet,contentDetails",
        maxResults=1,
        id=PLAYLIST_ID
    ).execute()

    playListTitle = response["items"][0]["snippet"]["title"]

    # request first set of playlist items
    response = youtube.playlistItems().list(
        part="snippet,contentDetails",
        maxResults=50,
        playlistId=PLAYLIST_ID
    ).execute()

    totalPages = math.ceil(response["pageInfo"]["totalResults"] / response["pageInfo"]["resultsPerPage"])
    if totalPages > MAX_PAGES:
        totalPages = MAX_PAGES

    print("Total pages to dump: " + str(totalPages))

    # request the rest of the pages of the playlist, and add their playlistItems resources into one big dictionary
    allItems = response["items"]

    token = ""
    if ("nextPageToken" in response):
        token = response["nextPageToken"] # token used to tell the API to return the next page

    # request all the other playlist items
    for x in range(totalPages - 1):
        response = youtube.playlistItems().list(
        part="snippet,contentDetails",
        maxResults=50,
        playlistId=PLAYLIST_ID,
        pageToken=token
        ).execute()

        for z in response["items"]:
            snippet = z["snippet"]
            title = snippet["title"]

            print("Adding " + title) # this is confusing, since we don't print for the first page of items

            allItems.append(z)

        if ("nextPageToken" in response): # update token to point to the next page for the next request
            token = response["nextPageToken"]

    # create a spreadsheet
    workbook = Workbook()
    sheet = workbook.active
    # sheet.title = response["items"][0]["snippet"]["title"] # does not work, it uses the item title
    sheet.title = playListTitle
    row = 2 # staring row. row 1 is for column info

    columnInfo = {
        "A": {
            "name": "Title",
            "size": 56
        },
        "B": {
            "name": "Author",
            "size": 18
        },
        "C": {
            "name": "Link",
            "size": 48
        },
        "D": {
            "name": "NND Link",
            "size": 40
        },
        "E": {
            "name": "Publish Time",
            "size": 24
        },
        "F": {
            "name": "Description",
            "size": 35
        },
    }

    # setup column widths.
    index = 1 # indexes start at 1 in spreadsheets
    for column in columnInfo:
        cell = sheet.cell(1, index)
        info = columnInfo[column]
        cell.value = info["name"]
        sheet.column_dimensions[column].width = info["size"]
        index += 1

    # iterate through items we received from the playlist items request
    for item in allItems:
        snippet = item["snippet"]

        title = snippet["title"]
        
        # don't try to check inaccessible videos
        if (title != "Deleted video" and title != "Private video"):

            publishTime = item["contentDetails"]["videoPublishedAt"]
            videoId = snippet["resourceId"]["videoId"]

            # get video info from the playlist item
            response = youtube.videos().list(
            part="snippet,contentDetails",
            id=videoId
            ).execute()

            video = response["items"][0]["snippet"] # [0] because it returns a list for some reason

            videoAuthor = video["channelTitle"]
            link = YT_VID_PREFIX + videoId
            description = video["description"].replace("\n", " ")
            descLines = description.splitlines() # regex doesnt work if we dont do this
            nndLink = ""

            # search for NND link in description
            for line in descLines:
                nnd = nndRegex.search(line)
                if (nnd != None):
                    nndLink = NND_VID_PREFIX + nnd.group(0)
                    print("Found NND link in video '" + title + "':")
                    print(nndLink)

        else: # todo check if we have access to any of this info from a request
            publishTime = "???"
            videoAuthor = "???"
            link = "???"
            nndLink = "???"
            description = "???"

        # this specifies the order in which info is put into each row
        info = [
            title,
            videoAuthor,
            link,
            nndLink,
            publishTime,
            description
        ]

        # iterate through cells in each row and input the data. 1 row = 1 playlist item
        for cell in range(1, len(info) + 1):
            x = sheet.cell(row, cell)
            x.value = info[cell - 1]

        row += 1

    # save the spreadsheet
    try:
        workbook.save(filename="dump.xlsx")
    except PermissionError:
        print("Make sure you don't have the sheet open!")

if __name__ == "__main__":
    main()